'''
Emma Schumacher
Katzlab
Spring 2022

Code to visualize germline some sequences, main code
Should be used together with makefigure.py & seqclass.py
Should produce a pngs & a spreadsheet called 'datasummary.xls' with 2 sheets
'''

import logging  # log file of data that needs to be looked at
import pandas as pd  # used to convert a .tsv
import sys  # take commandline arguments
import os  # to take deal with extra file
import xlwt  # to write to spreadsheet
from xlwt import Workbook  # to write to spreadsheet
from openpyxl import load_workbook  # to read from spreadsheet
import seqclassms  # class script to organize data
import makefigms  # script that makes the actual graph
 

def main():
    # so i dont get a million obnoxious matplotlib notes in my log
    logging.getLogger('matplotlib.font_manager').disabled = True

    # takes filename as command line argument
    fn = sys.argv[1]
    converted = False

    # since I realized too late Wumei's files are tsvs, handy built in converter
    if (fn.endswith('.tsv')):
        fn = handyconverter(fn)
        converted = True

    # Workbooks are created
    wb = Workbook()
    rb = load_workbook(filename=fn)  # source of data
    r_sheet = rb.active  # analysis spreadsheet

    # add_sheet is used to create sheets for analyzed data
    w_sheet = wb.add_sheet('Transcripts data')
    w_sheet2 = wb.add_sheet('Segments data')

    # Gets LKH
    lkh = str(fn[8:-5])
    print(lkh)

    # Makes a folder to store stuff in
    if not os.path.isdir(lkh):  # checks there doesnt already exist a folder for this LKH
        os.mkdir(lkh)
    os.chdir(lkh)  # changes directory  to new one (all output goes there)

    # makes log file of data that is sus and shoudl be double checked
    logging.basicConfig(filename='sus4' + lkh + '.log',
                        filemode='w', level=logging.DEBUG)

    # initial categories for worksheet on whole transcript x germ pairings
    wsheetinit(w_sheet2, 'S start', 'S end',
               'G start', 'G end', 'S GAP LEN', 'G GAP LEN', 'FLIPPED GERMLINE')
    
    wsheetinit(w_sheet, '% MAPPED', 'AVE LENGTH',
               '# SEGMENTS', 'MAPPED?', 'SOMA SCRAMBLED', 'GERMLINE SCRAMBLED', 'S LENGTH')

    # initialize some variables
    trandict = {}  # empty dictionary, key is transcript name value is list of sequence objects
    trnscrpt = "null"  # starter value for name to check against so I dont have duplicates
    gtrnscrpt = "null"  # starter value for name to check against so I dont have duplicates
    seg = 1  # number of segments in each transcript
    maplen = 0  # total length of sequences mapped (germline)
    rownum = 0  # number of transcripts total
    rownum2 = 0  # number of sequences
    slist = []
    glist = []

    # traverse spreadsheet by rows (skips categories on row 1)
    for r in range(2, r_sheet.max_row + 1):
        # get transcript ID & germline ID
        name = r_sheet.cell(row=r, column=1).value
        gname = r_sheet.cell(row=r, column=2).value

        maplen = 0
        smaplen = 0
        # checks if this is a new sequence
        if (trnscrpt not in name):
            # if not the first transcript
            if (trnscrpt != "null"):
                if (glist[0].get_rev() != slist[0].get_rev()):
                    for trnscrpt in glist:
                        trnscrpt.set_flip(True) #set all flips to true ??? 
                        
                        #set all revs to opposite ???
                        if (trnscrpt.get_rev() == True): 
                            trnscrpt.set_rev(False) 
                        else:
                            trnscrpt.set_rev(True)
                             
                # sorts lists
                glist.sort(key=lambda x: x.start)
                slist.sort(key=lambda x: x.start)

                #fills out worksheet now that it's sorted
                for i in range (0, len(glist)):
                    # fills out the 'Segments data' sheet in spreadsheet ???
                    rownum2 = rownum2 + 1  # moves down a row
                    
                    # calls wsheet2 function to write it up
                    if (i != 0):
                        #sets up gap between current and previous segment
                        slist[i].set_gaplen(slist[i].get_start() - slist[i-1].get_end()) #???
                        glist[i].set_gaplen(glist[i].get_start() - glist[i-1].get_end())
                        #sets up writing
                        a, b, c, d = sgmntwsheet(slist[i], glist[i], w_sheet2, rownum2, d, slist[i - 1].get_end(), glist[i - 1].get_end())
                    else:
                        d = 0 #WHAT IS D???
                        slist[i].set_gaplen(0) #???
                        glist[i].set_gaplen(0)
                        a, b, c, d = sgmntwsheet(slist[i], glist[i], w_sheet2, rownum2, d)

                    # initializes a bunch of variables
                    if (a == True):
                        srev = True
                    if (b == True):
                        grev = True
                    if (c == True):
                        somagap = True

                    # checks scrambling ???
                    if (i != 0):
                        if (slist[i - 2].get_rev() != a):
                            sscram = True
                    if (i != 0):
                        if (glist[i - 2].get_rev() != b):
                            gscram = True
                         
                
                # adds lists
                seglist.append(glist)
                seglist.append(slist)

                # Fills out 'Transcripts data' for the sequence that just ended 
                if (trnscrptwsheet(seglist, w_sheet, rownum, seg, srev, grev, somagap, sscram, gscram, d)):
                    # adds new dictionary entry with transcript contents if it was worth mapping
                    trandict[trnscrpt] = seglist
                 
                
            # until the last segment is being graphed
            if (r < r_sheet.max_row):
                # resets important variables and sets up for next transcript
                rownum = rownum + 1  # moves down to a new row in datasummary worksheets
                seg = 1  # resets counter for the number of segments in the transcript
                slist = []
                glist = []
                seglist = []  # makes an empty list of segments
                srev = False  # resets reversal checker
                grev = False  # resets reversal checker
                sscram = False  # resets scrambling checker
                gscram = False  # resets scrambling checker
                somagap = False  # resets irregular soma gaps checker
                trnscrpt = name[name.index("Tran"):name.index(
                    "_Len")]  # gets new transcript name
                gtrnscrpt = gname[gname.index("Germ"):gname.index(
                    "_Len")]  # gets new germ name

        else:  # if segment is not from a new transcript/germline
            seg = seg + 1  # add one sequence to counter

        # Creates sequence object for this row ???
        s = seqclassms.Soma(trnscrpt, r_sheet.cell(
            row=r, column=5).value, r_sheet.cell(row=r, column=6).value,
            0, 0)
        g = seqclassms.Germline(gtrnscrpt, r_sheet.cell(
            row=r, column=7).value, r_sheet.cell(row=r, column=8).value,
            0, 0, gtrnscrpt)
        s.check_rev()  # checks for soma reversal (affects object attributes/summary sheet)
        g.check_rev()  # checks for germline reversal (affects object attributes/summary sheet)
        s.set_size()  # sets size of MSG
        g.set_size()  # sets size of MSG

        # adds the current segment object to list for this transcriptxgerm
        slist.append(s)
        glist.append(g)

    # if i had a .tsv that i converted, deletes the xslx copy to save space ???
    if (converted == True):
        os.remove(fn)

    # closes open workbook
    rb.close()

    # saves workbook
    wb.save(lkh + 'datasummary.xls')

    # calls makefig to graph each of the transcriptsxgermlines that met my conditions
    for key in trandict:
        makefigms.graph(trandict[key])

# some titles I reuse


def wsheetinit(wsheets, col6, col7, col8, col9, col10, col11, col12):
    wsheets.write(0, 0, 'TRANSCRIPT')
    wsheets.write(0, 1, 'GERMLINE')
    wsheets.write(0, 2, 'TOTAL LENGTH')
    wsheets.write(0, 3, 'IRREGULAR SOMA GAP')
    wsheets.write(0, 4, 'SOMA REVERSAL')
    # ??? also need one for scrambling
    wsheets.write(0, 5, 'GERMLINE REVERSAL')

    # variable titles
    wsheets.write(0, 6, col6)
    wsheets.write(0, 7, col7)
    wsheets.write(0, 8, col8)
    wsheets.write(0, 9, col9)
    wsheets.write(0, 10, col10)
    wsheets.write(0, 11, col11)
    wsheets.write(0, 12, col12)


# this function writes info from the transcript object into my 'Segments data' sheet in spreadsheet
def sgmntwsheet(stranscript, gtranscript, wsheets, rownum, total, prevSend=0, prevGend=0, prevSstart=0, prevGstart=0):

    # writes the name of the old transcript
    wsheets.write(rownum, 0, stranscript.get_name())

    # writes the name of the old germline
    wsheets.write(rownum, 1, gtranscript.get_germ())

    # writes total length mapped
    wsheets.write(rownum, 2, gtranscript.get_length())

    # sets soma gap ???
    if (prevSend != 0 | prevGend != 0):  # checks gap before this sequence, so doesnt apply to first MDS ???
        stranscript.get_gaplen()  # gets soma gao
        gtranscript.get_gaplen()  # sets gaplength ???

        wsheets.write(rownum, 10, stranscript.get_gaplen())  # writes soma gap length
        wsheets.write(rownum, 11, gtranscript.get_gaplen())  # writes soma gap length
    else:
        wsheets.write(rownum, 10, "N/A")  # writes not applicable ???
        wsheets.write(rownum, 11, "N/A")  # writes not applicable ???

    if (stranscript.get_somagap() == False):
        wsheets.write(rownum, 3, "NO")  # indicates there is no somagap
        gap = False
    else:
        wsheets.write(rownum, 3, "YES")  # indicates there is an somagap
        gap = True

    if (stranscript.get_rev() == False):
        wsheets.write(rownum, 4, "NO")  # indicates there is no soma scrambling
        s = False
    else:
        wsheets.write(rownum, 4, "YES")  # indicates there is some scrambling
        s = True

    if (gtranscript.get_rev() == False):
        # indicates there is no germline scrambling
        wsheets.write(rownum, 5, "NO")
        g = False
    else:
        # indicates there is germline scrambling
        wsheets.write(rownum, 5, "YES")
        g = True

    if (gtranscript.get_flip() == False):
        # indicates there is no germline flip
        wsheets.write(rownum, 13, "NO")
        g = False
    else:
        # indicates there is germline flip
        wsheets.write(rownum, 13, "YES")
        g = True

    wsheets.write(rownum, 6, stranscript.get_start())
    wsheets.write(rownum, 7, stranscript.get_end())
    wsheets.write(rownum, 8, gtranscript.get_start())
    wsheets.write(rownum, 9, gtranscript.get_end())

    total = total + gtranscript.get_gaplen() + gtranscript.get_size()
    return (s, g, gap, total)

# this function writes info from the transcript sequences into my 'Transcript data' sheet in spreadsheet


def trnscrptwsheet(seglist, wsheets, rownum, seg, srev, grev, gap, sscram, gscram, d):
    maplen = 0
    smaplen = 0
    temp = 0
    propmap = 0
    total = 0

    # writes the name of the old transcript
    wsheets.write(rownum, 0, seglist[len(seglist)-1][0].get_name())

    wsheets.write(rownum, 1, seglist[0][0].get_germ())

    for i in range(len(seglist[0])):  # gets segments
        total += seglist[0][i].get_length()  # adds length to total (used for average seg)

        if (i < (len(seglist[0]) - 1)):  # if not last segment
            maplen += (seglist[0][i].get_length() + seglist[0][i].get_gaplen())#seglist[0][i].get_overlap(seglist[0][i+1].get_start()))
            smaplen += (seglist[1][i].get_size() + seglist[1][i].get_gaplen()) #???
            # writes to logmap if pointer >16 bp (# from Xyrus???)
            if (seglist[0][i].get_overlap(seglist[0][i+1].get_start()) > 16):
                logging.info('Further investigate: ' +
                             seglist[1][0].get_name() + 'x' + seglist[0][0].get_name())
                logging.warning("Weirdly long pointer on Germline (>16 bp) with length: " +
                                str(seglist[0][i].get_overlap(seglist[0][i+1].get_start())))
                logging.info('-' * 80)
        else:
            maplen += (seglist[0][i].get_length())

    wsheets.write(rownum, 2, maplen)  # writes total length mapped
    wsheets.write(rownum, 12, smaplen)

    if (gap == False):
        wsheets.write(rownum, 3, "NO")  # indicates there is no somagap
    else:
        wsheets.write(rownum, 3, "YES")  # indicates there is an somagap

    if (srev == False):
        wsheets.write(rownum, 4, "NO")  # indicates there is no soma reversal
    else:
        wsheets.write(rownum, 4, "YES")  # indicates there is some reversal

    if (grev == False):
        # indicates there is no germline reversal
        wsheets.write(rownum, 5, "NO")
    else:
        wsheets.write(rownum, 5, "YES")  # indicates there is germline reversal

    if (sscram == False):
        # indicates there is no soma scrambling
        wsheets.write(rownum, 10, "NO")
    else:
        wsheets.write(rownum, 10, "YES")  # indicates there is some scrambling

    if (gscram == False):
        # indicates there is no germline scrambling
        wsheets.write(rownum, 11, "NO")
    else:
        # indicates there is germline scrambling
        wsheets.write(rownum, 11, "YES")

    # calculates proportion mapped

    # difference between first & last points on germline ???
    propmap = round(100.0 * maplen/float(d))

    wsheets.write(rownum, 6, propmap)  # writes total proportion mapped

    # writes total proportion mapped
    wsheets.write(rownum, 7, round(total/float(seg)))

    wsheets.write(rownum, 8, seg)

    # it is worth graphing (maps to over 80% of germline, more than 2 segments, at least 100 bp) ???
    if ((propmap >= 80) & (seg > 2) & (maplen >= 100)):
        wsheets.write(rownum, 9, "YES")  # indicates that it will be graphed
        #g_scale = float((maplen)/(smaplen)) #??? 
        return (True)

        # it is not worth graphing (segment sucks)
    else:
        wsheets.write(rownum, 9, "NO")  # indicates it will not be graphed
        # explains why this sequence is garbage (prints to terminal)
        print((seglist[len(seglist)-1][0].get_name()) +
              " was not worth graphing because it did not:")
        if (propmap < 80):
            print(" - Map to at least 80% of the rDNA")
        if (seg <= 2):
            print(" - Have more than 2 segments")
        if (maplen < 100):
            print(" - Map a total length of at least 100 base pairs")
        print("")



# this function converts a .tsv int a .xlsx
def handyconverter(fn):
    # fixes name
    xlsx = str(fn[0:len(fn) - 4]) + '.xlsx'
    # reads tsv
    data = pd.read_csv(fn, sep='\t')
    # converts
    data.to_excel(xlsx, index=None, header=True)

    # return new filename
    return (xlsx)
 


# main function for coding hygeine
if __name__ == "__main__":
    main()
